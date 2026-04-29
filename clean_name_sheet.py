import openpyxl
import re

file_path = 'excel_workspace/항공권 정산/항공권 결제내역 정산(04월분)_최종_정밀완성_v2.xlsx'
output_file = 'excel_workspace/항공권 정산/항공권 결제내역 정산(04월분)_최종_최종_정합성완료.xlsx'

try:
    print("[영문국문명] 시트 데이터 정제 시작...")
    wb = openpyxl.load_workbook(file_path)
    
    # 1. [영문국문명] 시트 찾기
    try:
        ws_name_map = next(s for s in wb.worksheets if '영문국문명' in s.title)
        print(f"시트 발견: {ws_name_map.title}")
        
        count = 0
        # B열(2번째 열)의 공백 제거
        for row in range(2, ws_name_map.max_row + 1):
            cell = ws_name_map.cell(row=row, column=2)
            if cell.value and isinstance(cell.value, str):
                original_val = cell.value
                # 모든 공백 제거
                cleaned_val = original_val.replace(" ", "")
                if original_val != cleaned_val:
                    cell.value = cleaned_val
                    count += 1
        print(f"총 {count}개의 영문명에서 공백을 제거했습니다.")
    except StopIteration:
        print("경고: [영문국문명] 시트를 찾을 수 없습니다.")

    # 2. 저장
    wb.save(output_file)
    print(f"\n최종 정제 완료! 파일이 생성되었습니다: {output_file}")

except Exception as e:
    print(f"오류 발생: {e}")
