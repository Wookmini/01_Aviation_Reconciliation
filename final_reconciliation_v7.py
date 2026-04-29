import openpyxl
from openpyxl.styles import PatternFill
import re

sample_file = 'excel_workspace/항공권 정산/항공권 결제내역 정산(04월분)_sample.xlsx'
output_file = 'excel_workspace/항공권 정산/항공권 결제내역 정산(04월분)_최종_정밀완성_v7.xlsx'

def normalize_no(val):
    if val is None: return ""
    s = re.sub(r'[^0-9]', '', str(val))
    return s.lstrip('0')

def clean_amt(val):
    if val is None: return 0
    if isinstance(val, (int, float)): return float(val)
    s = str(val).replace('KRW', '').replace(',', '').strip()
    try:
        s = ''.join(c for c in s if c.isdigit() or c == '.' or c == '-')
        return float(s) if s else 0
    except:
        return 0

try:
    print("v7 정밀 금액 대조 작업을 시작합니다...")
    wb = openpyxl.load_workbook(sample_file)
    ws_dahs = next(s for s in wb.worksheets if 'e-DAHS' in s.title)
    ws_sejung = next(s for s in wb.worksheets if '세중' in s.title)
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # 1. e-DAHS 데이터 맵핑 (승인번호 -> 합계금액U(21열))
    dahs_data = {} # {norm_no: total_amt}
    dahs_approval_numbers = set()
    
    for row in range(2, ws_dahs.max_row + 1):
        raw_no = ws_dahs.cell(row=row, column=4).value # D열
        norm_no = normalize_no(raw_no)
        total_amt = clean_amt(ws_dahs.cell(row=row, column=21).value) # U열
        
        if norm_no:
            dahs_data[norm_no] = total_amt
            dahs_approval_numbers.add(norm_no)

    # 2. 세중 발권리스트 분석 및 색칠
    for row in range(2, ws_sejung.max_row + 1):
        ac_val = normalize_no(ws_sejung.cell(row=row, column=29).value) # AC열
        av_val = normalize_no(ws_sejung.cell(row=row, column=48).value) # AV열
        
        ah_amt = clean_amt(ws_sejung.cell(row=row, column=34).value) # AH열 (Net Amt)
        at_amt = clean_amt(ws_sejung.cell(row=row, column=46).value) # AT열 (취급수수료총액)
        
        is_error = True # 기본값은 오류(빨간색)
        is_missing = False # 존재 여부(노란색)

        # (1) 승인번호(AC) 매칭 및 금액(AH vs U) 검증
        if ac_val in dahs_data:
            if abs(ah_amt - dahs_data[ac_val]) < 1:
                is_error = False # 정상 매칭
            else:
                is_error = True # 금액 불일치
        
        # (2) 취급승인번호(AV) 매칭 및 금액(AT vs U) 검증
        # AC에서 정상이 아니었을 때만 AV 체크 (혹은 둘 다 체크)
        if is_error and av_val in dahs_data:
            if abs(at_amt - dahs_data[av_val]) < 1:
                is_error = False # 정상 매칭
            else:
                is_error = True # 금액 불일치

        # 매칭 정보가 아예 없는지 확인 (기존 Yellow 조건)
        if not (ac_val in dahs_approval_numbers or av_val in dahs_approval_numbers):
            is_missing = True

        # 배경색 적용
        if is_missing:
            # 아예 없으면 노란색 (v6 유지)
            for col in range(1, ws_sejung.max_column + 1):
                ws_sejung.cell(row=row, column=col).fill = yellow_fill
        elif is_error:
            # 번호는 있는데 금액이 틀리면 빨간색 (v7 신규)
            for col in range(1, ws_sejung.max_column + 1):
                ws_sejung.cell(row=row, column=col).fill = red_fill

    # 3. e-DAHS 데이터 기입 (기존 v6 로직 유지)
    # (세중 리스트를 다시 훑어 룩업 생성)
    sejung_lookup = {}
    for row in range(2, ws_sejung.max_row + 1):
        ac = normalize_no(ws_sejung.cell(row=row, column=29).value)
        av = normalize_no(ws_sejung.cell(row=row, column=48).value)
        info = (ws_sejung.cell(row=row, column=12).value, ws_sejung.cell(row=row, column=19).value, ws_sejung.cell(row=row, column=20).value)
        if ac: sejung_lookup[ac] = info
        if av: sejung_lookup[av] = info

    for row in range(2, ws_dahs.max_row + 1):
        norm_no = normalize_no(ws_dahs.cell(row=row, column=4).value)
        if norm_no in sejung_lookup:
            val = sejung_lookup[norm_no]
            ws_dahs.cell(row=row, column=5).value = val[0] # E: 영문명
            ws_dahs.cell(row=row, column=9).value = val[1] # I: 출발일
            ws_dahs.cell(row=row, column=10).value = val[2] # J: 여정

    wb.save(output_file)
    print(f"\n[v7 업데이트 완료]")
    print(f"- 금액 대조 및 오류(Red) 표시 완료")
    print(f"- 최종 파일: {output_file}")

except Exception as e:
    print(f"오류 발생: {e}")
