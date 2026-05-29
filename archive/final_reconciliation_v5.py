import openpyxl
from openpyxl.styles import PatternFill

# 파일 경로
sample_file = 'excel_workspace/항공권 정산/항공권 결제내역 정산(04월분)_sample.xlsx'
output_file = 'excel_workspace/항공권 정산/항공권 결제내역 정산(04월분)_최종_정밀완성.xlsx'

try:
    print("정밀 정산 작업을 시작합니다...")
    wb = openpyxl.load_workbook(sample_file)
    
    # 시트 이름 유연하게 찾기 (공백 등 대응)
    ws_dahs = next(s for s in wb.worksheets if 'e-DAHS' in s.title)
    ws_sejung = next(s for s in wb.worksheets if '세중' in s.title)
    
    print(f"시트 로드 완료: {ws_dahs.title}, {ws_sejung.title}")

    # 1. 세중 발권리스트 데이터 인덱싱 (AC=29번째, AV=48번째 열, 1-based)
    # 룩업 테이블 생성: {승인번호: (고객명영문, 출발일, 도시)}
    sejung_lookup = {}
    sejung_approval_numbers = set()
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in range(2, ws_sejung.max_row + 1):
        ac_val = str(ws_sejung.cell(row=row, column=29).value).strip() if ws_sejung.cell(row=row, column=29).value else None
        av_val = str(ws_sejung.cell(row=row, column=48).value).strip() if ws_sejung.cell(row=row, column=48).value else None
        
        eng_name = ws_sejung.cell(row=row, column=12).value # L열
        start_date = ws_sejung.cell(row=row, column=19).value # S열
        city = ws_sejung.cell(row=row, column=20).value # T열

        if ac_val:
            sejung_lookup[ac_val] = (eng_name, start_date, city)
            sejung_approval_numbers.add(ac_val)
        if av_val:
            sejung_lookup[av_val] = (eng_name, start_date, city)
            sejung_approval_numbers.add(av_val)

    # 2. e-DAHS 전표내역 업데이트
    dahs_approval_numbers = set()
    for row in range(2, ws_dahs.max_row + 1):
        approval_no = str(ws_dahs.cell(row=row, column=4).value).strip() if ws_dahs.cell(row=row, column=4).value else None
        if approval_no:
            dahs_approval_numbers.add(approval_no)
            if approval_no in sejung_lookup:
                val = sejung_lookup[approval_no]
                ws_dahs.cell(row=row, column=5).value = val[0] # E열: 이용자명(영문)
                ws_dahs.cell(row=row, column=9).value = val[1] # I열: 출발일자
                ws_dahs.cell(row=row, column=10).value = val[2] # J열: 여정
    
    # 3. 세중 발권리스트 미매칭 건 노란색 배경색 칠하기
    for row in range(2, ws_sejung.max_row + 1):
        ac_val = str(ws_sejung.cell(row=row, column=29).value).strip() if ws_sejung.cell(row=row, column=29).value else None
        av_val = str(ws_sejung.cell(row=row, column=48).value).strip() if ws_sejung.cell(row=row, column=48).value else None
        
        # AC 또는 AV 번호가 e-DAHS에 없으면 노란색 칠하기
        is_matched = (ac_val in dahs_approval_numbers) or (av_val in dahs_approval_numbers)
        if not is_matched and (ac_val or av_val):
            for col in range(1, ws_sejung.max_column + 1):
                ws_sejung.cell(row=row, column=col).fill = yellow_fill

    # 4. 저장
    wb.save(output_file)
    print(f"정밀 정산 완료! 최종 파일: {output_file}")

except Exception as e:
    print(f"오류 발생: {e}")
    import traceback
    traceback.print_exc()
