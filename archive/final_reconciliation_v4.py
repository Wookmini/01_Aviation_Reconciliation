import openpyxl
from openpyxl.styles import PatternFill

def reconcile_airline_tickets():
    input_file = './excel_workspace/항공권 정산/항공권 결제내역 정산(04월분)_sample.xlsx'
    output_file = './excel_workspace/항공권 정산/항공권 결제내역 정산(04월분)_최종_정밀완성.xlsx'

    print(f"Loading workbook: {input_file}")
    wb = openpyxl.load_workbook(input_file, data_only=False) # Keep formulas
    
    sheet_names = wb.sheetnames
    print(f"Sheets: {sheet_names}")

    dahs_sheet_name = next((s for s in sheet_names if 'e-DAHS' in s), None)
    sejung_sheet_name = next((s for s in sheet_names if '세중' in s or '諛沅由ъㅽ' in s), None) # Handle possible garbled name
    settlement_sheet_name = next((s for s in sheet_names if '정산내역' in s or '寃곗댁' in s), None)

    if not dahs_sheet_name or not sejung_sheet_name:
        # Fallback if names are completely different
        print("Warning: Could not find sheets by keyword. Using indexes if available.")
        # Let's try to be more specific if possible. 
        # From previous output: 'e-DAHS  댁' was index 2, ' 몄 諛沅由ъㅽ' was index 3
        if len(sheet_names) > 3:
            dahs_sheet_name = sheet_names[2]
            sejung_sheet_name = sheet_names[3]
        else:
            print("Error: Required sheets not found.")
            return

    print(f"Using Dahs sheet: {dahs_sheet_name}")
    print(f"Using Sejung sheet: {sejung_sheet_name}")

    dahs_ws = wb[dahs_sheet_name]
    sejung_ws = wb[sejung_sheet_name]
    
    # 1. Map Sejung data for quick lookup by approval number (AC=29, AV=48)
    sejung_map = {}
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row_idx in range(2, sejung_ws.max_row + 1):
        ac_val = sejung_ws.cell(row=row_idx, column=29).value
        av_val = sejung_ws.cell(row=row_idx, column=48).value
        
        data = {
            'L': sejung_ws.cell(row=row_idx, column=12).value,
            'S': sejung_ws.cell(row=row_idx, column=19).value,
            'T': sejung_ws.cell(row=row_idx, column=20).value
        }
        
        if ac_val:
            sejung_map[str(ac_val).strip()] = data
        if av_val:
            sejung_map[str(av_val).strip()] = data

    # 2. Update e-DAHS sheet (D=4, E=5, I=9, J=10)
    dahs_approval_numbers = set()
    for row_idx in range(2, dahs_ws.max_row + 1):
        d_val = dahs_ws.cell(row=row_idx, column=4).value
        if not d_val:
            continue
        
        d_str = str(d_val).strip()
        dahs_approval_numbers.add(d_str)
        
        if d_str in sejung_map:
            match_data = sejung_map[d_str]
            dahs_ws.cell(row=row_idx, column=5).value = match_data['L'] # E
            dahs_ws.cell(row=row_idx, column=9).value = match_data['S'] # I
            dahs_ws.cell(row=row_idx, column=10).value = match_data['T'] # J (Corrected key to 'T')
            # F, G, H are left as is (formulas)

    # 3. Highlight rows in Sejung list not found in e-DAHS
    for row_idx in range(2, sejung_ws.max_row + 1):
        ac_val = sejung_ws.cell(row=row_idx, column=29).value
        av_val = sejung_ws.cell(row=row_idx, column=48).value
        
        ac_str = str(ac_val).strip() if ac_val else None
        av_str = str(av_val).strip() if av_val else None
        
        found = False
        if ac_str and ac_str in dahs_approval_numbers:
            found = True
        if not found and av_str and av_str in dahs_approval_numbers:
            found = True
            
        if not found:
            if ac_val or av_val: # Only highlight if there's actually an approval number to check
                for col_idx in range(1, sejung_ws.max_column + 1):
                    sejung_ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

    # 4. Save the result
    wb.save(output_file)
    print(f"Reconciliation completed and saved to: {output_file}")

if __name__ == "__main__":
    reconcile_airline_tickets()
