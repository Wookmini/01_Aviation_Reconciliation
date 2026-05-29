import openpyxl
from openpyxl.styles import PatternFill
import re

sample_file = 'excel_workspace/항공권 정산/항공권 결제내역 정산(04월분)_sample.xlsx'
output_file = 'excel_workspace/항공권 정산/항공권 결제내역 정산(04월분)_최종_정밀완성_v2.xlsx'

def normalize_no(val):
    if val is None: return ""
    # 숫자만 남기고 모두 제거
    s = re.sub(r'[^0-9]', '', str(val))
    # 앞의 0 제거 (02607603 -> 2607603 처리)
    return s.lstrip('0')

try:
    print("유연한 승인번호 매칭 작업을 시작합니다...")
    wb = openpyxl.load_workbook(sample_file)
    ws_dahs = next(s for s in wb.worksheets if 'e-DAHS' in s.title)
    ws_sejung = next(s for s in wb.worksheets if '세중' in s.title)
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 1. 세중 발권리스트 인덱싱 (정규화된 번호 사용)
    sejung_lookup = {}
    sejung_norm_numbers = set()
    
    for row in range(2, ws_sejung.max_row + 1):
        ac_val = normalize_no(ws_sejung.cell(row=row, column=29).value)
        av_val = normalize_no(ws_sejung.cell(row=row, column=48).value)
        
        eng_name = ws_sejung.cell(row=row, column=12).value
        start_date = ws_sejung.cell(row=row, column=19).value
        city = ws_sejung.cell(row=row, column=20).value

        if ac_val:
            sejung_lookup[ac_val] = (eng_name, start_date, city)
            sejung_norm_numbers.add(ac_val)
        if av_val:
            sejung_lookup[av_val] = (eng_name, start_date, city)
            sejung_norm_numbers.add(av_val)

    # 2. e-DAHS 업데이트 (정규화된 번호 대조)
    dahs_norm_numbers = set()
    update_count = 0
    
    for row in range(2, ws_dahs.max_row + 1):
        raw_no = ws_dahs.cell(row=row, column=4).value
        norm_no = normalize_no(raw_no)
        
        if norm_no:
            dahs_norm_numbers.add(norm_no)
            if norm_no in sejung_lookup:
                val = sejung_lookup[norm_no]
                ws_dahs.cell(row=row, column=5).value = val[0]
                ws_dahs.cell(row=row, column=9).value = val[1]
                ws_dahs.cell(row=row, column=10).value = val[2]
                update_count += 1
                if 50 <= row <= 62:
                    print(f"Row {row} 매칭 성공! (번호: {norm_no})")

    # 3. 세중 미매칭 건 하이라이트
    for row in range(2, ws_sejung.max_row + 1):
        ac_val = normalize_no(ws_sejung.cell(row=row, column=29).value)
        av_val = normalize_no(ws_sejung.cell(row=row, column=48).value)
        
        is_matched = (ac_val in dahs_norm_numbers) or (av_val in dahs_norm_numbers)
        if not is_matched and (ac_val or av_val):
            for col in range(1, ws_sejung.max_column + 1):
                ws_sejung.cell(row=row, column=col).fill = yellow_fill

    wb.save(output_file)
    print(f"\n작업 완료! 총 {update_count}건 업데이트됨.")
    print(f"최종 파일: {output_file}")

except Exception as e:
    print(f"오류 발생: {e}")
