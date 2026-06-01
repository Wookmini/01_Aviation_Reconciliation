import openpyxl
from openpyxl.styles import PatternFill
import re

sample_file = 'excel_workspace/항공권 정산/항공권 결제내역 정산_input.xlsx'
output_file = 'excel_workspace/항공권 정산/항공권 결제내역 정산_output.xlsx'

def normalize_no(val):
    if val is None: return ""
    s = re.sub(r'[^0-9]', '', str(val))
    return s.lstrip('0')

def clean_amt(val):
    if val is None: return 0
    if isinstance(val, (int, float)): return float(val)
    s = str(val).replace('KRW', '').replace(',', '').replace(' ', '').strip()
    try:
        s = ''.join(c for c in s if c.isdigit() or c == '.' or c == '-')
        return float(s) if s else 0
    except:
        return 0

def clean_name(val):
    if val is None: return ""
    return str(val).replace(" ", "").strip()

try:
    print("v9 동적 데이터 감지 정산 작업을 시작합니다...")
    wb = openpyxl.load_workbook(sample_file)
    
    # 1. [영문국문명] 시트 공백 제거 (B열)
    try:
        ws_name_map = next(s for s in wb.worksheets if '영문국문명' in s.title)
        for row in range(2, ws_name_map.max_row + 1):
            cell = ws_name_map.cell(row=row, column=2)
            if cell.value:
                cell.value = clean_name(cell.value)
    except StopIteration:
        pass

    ws_dahs = next(s for s in wb.worksheets if 'e-DAHS' in s.title)
    ws_sejung = next(s for s in wb.worksheets if '세중' in s.title)
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # 2. e-DAHS 데이터 맵핑 (승인번호 -> 합계금액 U(21열))
    dahs_data = {}
    dahs_approval_numbers = set()
    
    for row in range(2, ws_dahs.max_row + 1):
        raw_no = ws_dahs.cell(row=row, column=4).value # D열
        norm_no = normalize_no(raw_no)
        total_amt = clean_amt(ws_dahs.cell(row=row, column=21).value) # U열
        if norm_no:
            dahs_data[norm_no] = total_amt
            dahs_approval_numbers.add(norm_no)

    # 3. 세중 발권리스트 동적 분석 및 색칠
    for row in range(2, ws_sejung.max_row + 1):
        ac_raw = ws_sejung.cell(row=row, column=29).value # AC열
        av_raw = ws_sejung.cell(row=row, column=48).value # AV열
        
        ac_val = normalize_no(ac_raw)
        av_val = normalize_no(av_raw)
        
        # [핵심] 승인번호와 취급승인번호가 모두 없으면 데이터가 없는 행으로 간주하고 스킵
        if not ac_val and not av_val:
            continue
            
        ah_amt = clean_amt(ws_sejung.cell(row=row, column=34).value) # AH열 (Net Amt)
        at_amt = clean_amt(ws_sejung.cell(row=row, column=46).value) # AT열 (취급수수료)
        
        is_matched = (ac_val in dahs_approval_numbers or av_val in dahs_approval_numbers)

        if not is_matched:
            # 번호 자체가 e-DAHS에 없으면 노란색
            for col in range(1, ws_sejung.max_column + 1):
                ws_sejung.cell(row=row, column=col).fill = yellow_fill
        else:
            # 번호는 있는데 금액이 틀리면 빨간색
            error_found = False
            if ac_val in dahs_data and abs(ah_amt - dahs_data[ac_val]) >= 1:
                error_found = True
            if not error_found and av_val in dahs_data and abs(at_amt - dahs_data[av_val]) >= 1:
                error_found = True
                
            if error_found:
                for col in range(1, ws_sejung.max_column + 1):
                    ws_sejung.cell(row=row, column=col).fill = red_fill

    # 4. e-DAHS 자동 기입 (v6 로직 기반)
    sejung_lookup = {}
    for row in range(2, ws_sejung.max_row + 1):
        ac = normalize_no(ws_sejung.cell(row=row, column=29).value)
        av = normalize_no(ws_sejung.cell(row=row, column=48).value)
        info = (ws_sejung.cell(row=row, column=12).value, ws_sejung.cell(row=row, column=19).value, ws_sejung.cell(row=row, column=20).value)
        if ac: sejung_lookup[ac] = info
        if av: sejung_lookup[av] = info

    for row in range(2, ws_dahs.max_row + 1):
        norm_no = normalize_no(ws_dahs.cell(row=row, column=4).value)
        if norm_no in sejung_lookup:
            val = sejung_lookup[norm_no]
            ws_dahs.cell(row=row, column=5).value = val[0] # E: 영문명
            ws_dahs.cell(row=row, column=9).value = val[1] # I: 출발일
            ws_dahs.cell(row=row, column=10).value = val[2] # J: 여정

    wb.save(output_file)
    print(f"\n[v9 업데이트 완료]")
    print("- 동적 데이터 영역 감지 로직 적용 (빈 행/요약 행 자동 스킵)")
    print("- 정산 및 하이라이트 기능 고도화 완료")
    print(f"- 최종 파일: {output_file}")

except Exception as e:
    print(f"오류 발생: {e}")
