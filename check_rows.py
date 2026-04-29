import openpyxl

file_path = 'excel_workspace/항공권 정산/항공권 결제내역 정산(04월분)_sample.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
ws_dahs = next(s for s in wb.worksheets if 'e-DAHS' in s.title)
ws_sejung = next(s for s in wb.worksheets if '세중' in s.title)

# 세중 승인번호 리스트업
sejung_nums = []
for row in range(2, ws_sejung.max_row + 1):
    ac = str(ws_sejung.cell(row=row, column=29).value).strip() if ws_sejung.cell(row=row, column=29).value else None
    av = str(ws_sejung.cell(row=row, column=48).value).strip() if ws_sejung.cell(row=row, column=48).value else None
    if ac: sejung_nums.append(ac)
    if av: sejung_nums.append(av)

print(f"--- 분석 시작 (행 50~62) ---")
for i in range(50, 63):
    approval_no = ws_dahs.cell(row=i, column=4).value
    str_no = str(approval_no).strip() if approval_no else "None"
    
    exists = str_no in sejung_nums
    print(f"Row {i}: [승인번호: {str_no}] -> 세중 리스트 존재 여부: {exists}")
    
    if exists:
        # 존재하는데 왜 안 채워졌는지 상세 확인
        print(f"  * 상세: 타입={type(approval_no)}")
